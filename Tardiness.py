import random
import itertools
import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
# Change directory to desktop
os.chdir(os.path.expanduser("~/Desktop"))


setup_times_df = pd.read_excel("433 setups.xlsx", index_col=0)
setup_times = setup_times_df.to_dict()


first_setup_times_df = pd.read_excel("Setup Times Full.xlsx", index_col=0, header=None, names=['Job', 'O'])
first_setup_times = first_setup_times_df['O'].to_dict()


processing_times_df = pd.read_excel("Processing Times Full.xlsx", index_col=0)
processing_times = processing_times_df.to_dict()

due_dates_df = pd.read_excel("Due Dates Full.xlsx", index_col=0, header=None, names=['Job', 'O'])
due_dates = due_dates_df['O'].to_dict()


from openpyxl import load_workbook

# Load the workbook
wb = load_workbook(filename='GA_Input_Output.xlsx')

# Select the sheet you want to read from
ws = wb['Data']

# Read values from cells
num_jobs = ws['B1'].value
num_machines = ws['B2'].value
population_size = ws['B3'].value
num_generations = ws['B4'].value
mutation_rate = ws['B5'].value
elite_percentage = 0.1
tournament_size = 5
random.seed(42)

def generate_individual():
    job_due_dates = [(job, due_dates[job]) for job in range(1, num_jobs+1)]
    sorted_jobs = sorted(job_due_dates, key=lambda x: x[1]) # sort the jobs by due date in ascending order
    machine_assignments = [0] * num_jobs # initialize machine assignments to 0 for all jobs
    machine_loads = [0] * num_machines # initialize machine loads to 0 for all machines

    for i, job in enumerate(sorted_jobs):
        min_load = min(machine_loads)  # find the lowest machine load
        min_load_machines = [i for i, load in enumerate(machine_loads) if load == min_load]  # find the machines with the lowest load
        machine_id = random.choice(min_load_machines)  # randomly select a machine from among those with the lowest load
        machine_assignments[job[0]-1] = machine_id + 1  # assign the job to the selected machine
        machine_loads[machine_id] += job[1]  # update the machine load with the due date of the assigned job

    return machine_assignments


def initial_population():
    population = []
    while len(population) < population_size:
        new_individual = generate_individual()
        if new_individual not in population:
            population.append(new_individual)
    return population



def total_tardiness(individual, debugg=0):
    machine_times = [0] * num_machines
    last_job_on_machine = [None] * num_machines
    machine_job_sequences = [[] for _ in range(num_machines)]

    tardiness = 0

    for job, machine in enumerate(individual, start=1):
        processing_time = processing_times[machine][job]
        setup_time = first_setup_times[job]

        if last_job_on_machine[machine - 1] is not None:
            prev_job = last_job_on_machine[machine - 1]
            setup_time = setup_times[prev_job][job]

        machine_times[machine - 1] += processing_time + setup_time
        last_job_on_machine[machine - 1] = job
        machine_job_sequences[machine - 1].append(job)

        completion_time = machine_times[machine - 1]
        tardiness += max(0, completion_time - due_dates[job])
        #if debugg == 1:
            #print(f"adding job {job} to machine {machine} processing time {processing_time} setup time {setup_time} completion time {completion_time}")

    return tardiness, machine_job_sequences, machine_times


# Selection: Tournament selection
def tournament_selection(population):
    return min(random.sample(population, tournament_size), key=lambda x: total_tardiness(x)[0])


# Crossover: One-point crossover
def crossover(parent1, parent2):
    crossover_point = random.randint(1, num_jobs - 1)
    child1 = parent1[:crossover_point] + parent2[crossover_point:]
    child2 = parent2[:crossover_point] + parent1[crossover_point:]
    return child1, child2


def mutate(individual):
    # Calculate the machine_job_sequences based on the individual
    _, machine_job_sequences, completion_times = total_tardiness(individual)

    # Find the machine with the longest completion time
    max_machine_idx = completion_times.index(max(completion_times))

    # Get the jobs on the machine with the longest completion time
    jobs_on_max_machine = machine_job_sequences[max_machine_idx]

    # Calculate the sum of setup time and processing time for each job on the machine
    job_times = [(job, processing_times[max_machine_idx + 1][job] + (setup_times[machine_job_sequences[max_machine_idx][index - 1]][job] if index > 0 else first_setup_times[job])) for index, job in enumerate(jobs_on_max_machine)]

    # Sort the jobs by the sum of setup time and processing time in descending order
    sorted_jobs = sorted(job_times, key=lambda x: x[1], reverse=True)

    # Try moving each job one by one to the machine with the shortest completion time except the current machine
    for job, _ in sorted_jobs:
        min_completion_time = float('inf')
        best_machine = -1
        for i, machine_completion_time in enumerate(completion_times):
            if i != max_machine_idx and machine_completion_time < min_completion_time:
                min_completion_time = machine_completion_time
                best_machine = i + 1

        # Update the individual with the new assignment
        original_individual = individual.copy()
        job_position = [index for index, machine in enumerate(individual) if machine == max_machine_idx + 1 and index + 1 == job]
        if job_position:
            individual[job_position[0]] = best_machine

        # If there's an improvement, stop
        if total_tardiness(individual)[0] < total_tardiness(original_individual)[0]:
            break
        else:
            individual = original_individual.copy()  # Revert the change

    return individual



def genetic_algorithm():
    population = initial_population()

    for generation in range(num_generations):
        new_population = []
        elites = sorted(population, key=lambda x: total_tardiness(x)[0])[:int(elite_percentage * population_size)]

        for _ in range(population_size // 2 - len(elites) // 2):
            parent1 = tournament_selection(population)
            parent2 = tournament_selection(population)

            child1, child2 = crossover(parent1, parent2)

            if random.random() < mutation_rate:
                mutated_child1 = mutate(child1)
                if total_tardiness(mutated_child1)[0] < total_tardiness(child1)[0]:
                    child1 = mutated_child1
            if random.random() < mutation_rate:
                mutated_child2 = mutate(child2)
                if total_tardiness(mutated_child2)[0] < total_tardiness(child2)[0]:
                    child2 = mutated_child2

            new_population += [child1, child2]

        population = elites + new_population

    return min(population, key=lambda x: total_tardiness(x)[0]), total_tardiness(min(population, key=lambda x: total_tardiness(x)[0]))[1]


best_solution, machine_job_sequences = genetic_algorithm()
print("Total tardiness:", total_tardiness(best_solution)[0])
print(total_tardiness(best_solution,1)[2])
print("Job orders on the machines:")
for i, sequence in enumerate(machine_job_sequences, start=1):
    print(f"Machine {i}: {sequence}")


import matplotlib.pyplot as plt

tardiness_start_times = {}
for machine_id, jobs in enumerate(machine_job_sequences, start=1):
    machine_start_time = 0
    for job in jobs:
        setup_time = first_setup_times[job]
        processing_time = processing_times[machine_id][job]

        if machine_job_sequences[machine_id - 1].index(job) > 0:
            prev_job = machine_job_sequences[machine_id - 1][machine_job_sequences[machine_id - 1].index(job) - 1]
            setup_time = setup_times[prev_job][job]

        tardiness_start_times[(job, machine_id)] = machine_start_time + setup_time
        machine_start_time += setup_time + processing_time

# Plot the Gantt chart for Genetic Algorithm
colors = plt.get_cmap('tab10').colors  # Use a colormap to support more colors
if total_tardiness(best_solution)[0] >= 0:
    fig, ax = plt.subplots()
    machines = []

    for k in range(1, num_machines + 1):
        for i in range(1, num_jobs + 1):
            if i in machine_job_sequences[k - 1]:
                machine_name = f"Machine {k}"
                start_time = tardiness_start_times[(i, k)]
                processing_time = processing_times[k][i]
                ax.broken_barh([(start_time, processing_time)], (k * 10, 9),
                               facecolors=(colors[i % len(colors)]))  # Use colors in a cyclic way
                machines.append(machine_name)
                ax.text(start_time + processing_time / 2, k * 10 + 4, f"Job {i}", ha='center', va='center', color='black')

    plt.title("Gantt chart (Total Tardiness Optimization)")
    plt.xlabel("Time")
    plt.yticks([k * 10 + 5 for k in range(1, num_machines + 1)], [f"Machine {k}" for k in range(1, num_machines + 1)])
    plt.grid(axis='x')
    ax.set_xlim(left=0)

    plt.show()

def calculate_completion_times(job_sequences, start_times):
    completion_times = []
    for machine_id, jobs in enumerate(job_sequences, start=1):
        machine_completion_times = []
        for job in jobs:
            processing_time = processing_times[machine_id][job]
            start_time = start_times[(job, machine_id)]
            machine_completion_times.append(start_time + processing_time)
        completion_times.append(machine_completion_times)
    return completion_times

completion_times = calculate_completion_times(machine_job_sequences, tardiness_start_times)

# Create a list to hold dataframes
df_list = []

# For each machine, create a dataframe and append it to the list
for i in range(num_machines):
    df_machine = pd.DataFrame({
        f'Machine_{i+1}': machine_job_sequences[i],
        f'Completion_Time_{i+1}': completion_times[i]
    })
    df_list.append(df_machine)

# Concatenate all dataframes horizontally
result_df = pd.concat(df_list, axis=1)

# Write the final dataframe to an Excel file
result_df.to_excel("Copy of Output Excel.xlsx", index=False)
